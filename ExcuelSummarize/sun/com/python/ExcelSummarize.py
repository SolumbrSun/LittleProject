import os
import openpyxl
import tkinter.messagebox

def write2Excel(FList, ExceList):
    FinalList = []
    writecase = openpyxl.Workbook()
    finalsheet = writecase.get_active_sheet()
    for i in range(len(FList[0])):
        finalsheet.cell(row = 1, column = i + 1).value = FList[0][i]
        FinalList.append(FList[0][i])
    for j in range(len(FList)):
        origintitle = openpyxl.load_workbook(ExceList[j])
        originsheet = origintitle.get_sheet_by_name('Sheet1')
        finalrow = finalsheet.max_row
        for k in range(len(FinalList)):
            # print(finalrow)
            for l in range(originsheet.max_row - 1):
                # print(finalrow+1)
                originrow = FList[j].index(FinalList[k])
                finalsheet.cell(row = finalrow + 1 + l  , column = k+1).value = originsheet.cell(row = 2+l, column = originrow + 1).value
    writecase.save('Final.xlsx')

def judgeTitle(ExceList):
    FList = []
    for i in range(len(ExceList)):
        CList = []
        titlecase = openpyxl.load_workbook(ExceList[i])
        sheet = titlecase.get_sheet_by_name('Sheet1')
        for j in range(sheet.max_column):
            CList.append(sheet.cell(row = 1, column = j + 1 ).value)
        FList.append(CList)
    # print(FList)
    for k in range(len(FList)):
        if len(FList[0]) == len(FList[k]):
            for l in range(len(FList[k])):
                if FList[k][l] not in FList[0]:
                    # print(str.format('第{0}个Excel的标题名称【{1}】与其它的存在差异，请检查所有Excel的标题名称是否一致！', k+1, FList[k][l]))
                    tkinter.messagebox.askokcancel('提示', str.format('第{0}个Excel的标题名称【{1}】与其它的存在差异，请检查所有Excel的标题名称是否一致！', k+1, FList[k][l]))
        else:
            # print(str.format('第{0}个Excel的标题数量与其它的不一致，请检查所有Excel的标题数量是否一致！', k+1))
            tkinter.messagebox.askokcancel('提示', str.format('第{0}个Excel的标题数量与其它的不一致，请检查所有Excel的标题数量是否一致！', k + 1))
    write2Excel(FList, ExceList)



def judgeExcel():
    global ExceList
    ExceList = []
    for foldername, subfolders, filenames in os.walk('.'):
        for filename in filenames:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                ExceList.append(filename)
    if len(ExceList) > 1:
        judgeTitle(ExceList)
    elif len(ExceList) == 1:
        # print('只存在一个Excel文件，不需要处理')
        tkinter.messagebox.askokcancel('提示', '只存在一个Excel文件，不需要处理')
    else:
        # print('不存在Excel文件，请将Excel路径下执行此程序')
        tkinter.messagebox.askokcancel('提示', '不存在Excel文件，请将Excel路径下执行此程序')



if __name__ == '__main__':
    root = tkinter.Tk()
    root.withdraw()
    if os.path.isfile('Final.xlsx'):
        os.unlink('Final.xlsx')
    judgeExcel()

























