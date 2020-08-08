from xmindparser import xmind_to_dict
import os
import openpyxl

# 获取xmind生成json
def getXmindJson(file):
    dict_out = xmind_to_dict(file)
    getMessage(dict_out)

# 创建Excel文件，如果存在则删除重建
def createExcel(Xmind_name):
    # 声明全局变量
    global Excel_name
    # 测试用例 替换 测试方案； xlsx 替换 xmind
    Excel_name = Xmind_name.replace('测试方案', '测试用例').replace('xmind', 'xlsx')
    if os.path.isfile(Excel_name):
        os.unlink(Excel_name)
    testcase = openpyxl.Workbook()
    sheet = testcase.get_active_sheet()
    sheet.title = '测试用例'
    title_row = ['序号', '需求名称', '功能模块', '用例名称', '用例描述', '优先级', '菜单目录', '前置条件', '输入/步骤', '预期结果', '实际结果', '状态', '用例执行人', '测试时间', '备注']
    # print(len(title_row))
    for i in range(len(title_row)):
        sheet.cell(row = 1, column = i + 1).value = title_row[i]
    testcase.save(Excel_name)

# 解析json
def getMessage(dict):
    # 获取文件名，用于生成Excel文件名
    Xmind_name = os.path.basename(xmind_file)
    createExcel(Xmind_name)
    # 获取功能模块的数量
    for i in range(len(dict[0]['topic']['topics'])):
        # 循环遍历，获取每个功能模块的测试用例数量
        for j in range(len(dict[0]['topic']['topics'][i]['topics'])):
            # 创建空列表，获取用例项
            Excel_list = []
            # 获取需求名称
            Excel_list.append(dict[0]['topic']['title'])
            # 循环遍历，获取功能模块，并拼接到需求名称后面
            Excel_list.append(dict[0]['topic']['topics'][i]['title'])
            # 获取每个测试用例的描述项（用例名称、用例描述、优先级、菜单目录、前置条件、输入/步骤、预期结果）
            for k in range(len(dict[0]['topic']['topics'][i]['topics'][j]['topics'])):
                Excel_list.append(dict[0]['topic']['topics'][i]['topics'][j]['topics'][k]['topics'][0]['title'])
            write2Excel(Excel_list)

def write2Excel(case_list):
    fillcase = openpyxl.load_workbook(Excel_name)
    sheet = fillcase.get_sheet_by_name('测试用例')
    case_row = sheet.max_row
    # print(case_row)
    for i in range(len(case_list)):
        sheet.cell(row = case_row + 1, column = 1).value = case_row
        sheet.cell(row=case_row + 1, column = i + 2).value = case_list[i]
    fillcase.save(Excel_name)

if __name__ == '__main__':
    xmind_file = "E:\\Python\\LittleProject\\Xmind2Excel\\sun\\com\\python\\国庆打劫小肥仔项目_测试方案_SolumbrSun.xmind"  # xmind文件
    getXmindJson(xmind_file)
